import openpyxl

# openpyxl.load_workbook('./../Data/Akshara.xlsx')
# value = wb['Sheet1'].cell(1, 1).value
# wb.close()
# print(value)

def get_xl_data(path,sheet,row,col):
    print('get data from xl')
    value=''
    try:
        wb=openpyxl.load_workbook(path)
        value = wb[sheet].cell(row, col).value
        wb.close()
        print('able to get the data from xl:', value)
    except:
        print('error while reading xl')
        value=''
    return value

print(get_xl_data('./../Data/Akshara.xlsx','Sheet1',1,1))